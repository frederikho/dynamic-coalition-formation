"""
Ingest precomputed payoff values from GDX files into a formatted Excel payoff table.

Reads `welfare_regional_t` (welfare per region per time period, summed up to a cutoff
year) and `W_SAI` (global SAI deployment level, also summed up to a cutoff year) from
each GDX file in a directory, where each file corresponds to one coalition state
(deployment scenario).  The period-to-year mapping is read from the `year` parameter
in the same GDX file.

State names in the output match the framework's coalition-structure notation exactly
(e.g. '( )', '(INDUSA)', '(INDRUSUSA)'), so the table can be consumed directly by
find_equilibrium with --payoff-table.

Writes a single formatted Excel file to payoff_tables/.

Usage:
    python -m lib.ingest_payoffs --input-dir /path/to/gdx/files
    python -m lib.ingest_payoffs --input-dir /path/to/gdx/files --output payoff_tables/burke.xlsx
    python -m lib.ingest_payoffs --cutoff-year 2100
    python -m lib.ingest_payoffs --help
"""

import argparse
import math
import re
import subprocess
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from itertools import chain, combinations

# Path to gdxdump binary
GDXDUMP = Path("/opt/gams/gams52.4_linux_x64_64_sfx/gdxdump")

# Welfare symbol to extract (one value per region per time period, GAMS parameter)
WELFARE_SYMBOL = "welfare_regional_t"

# SAI deployment symbol (one value per time period, GAMS variable → use .L levels)
SAI_SYMBOL = "W_SAI"

# Year mapping symbol (parameter: period index → calendar year)
YEAR_SYMBOL = "year"

# Base directory containing all RICE50x result folders
DEFAULT_RESULTS_DIR = Path("/home/frederik/Code/RICE50x/results")

# Default input directory containing GDX files (used when running ingest_payoffs directly)
DEFAULT_INPUT_DIR = DEFAULT_RESULTS_DIR / "burke"

# Default cutoff year for W_SAI summation
DEFAULT_CUTOFF_YEAR = 2060

# Default players: (gdx_region_code, display_name)
# Display names are the RICE50x codes uppercased (e.g. 'nde' → 'NDE').
# See lib/rice50x_regions.py for the full list of region codes and country names.
DEFAULT_PLAYERS = [
    ("nde", "NDE"),
    ("usa", "USA"),
    ("rus", "RUS"),
]

# Colors matching existing strategy table style
COLORS = {
    "title":      "FF4D9CC9",  # Blue
    "header":     "FFD3D3D9",  # Light gray
    "state_name": "FFEEBF99",  # Beige/orange
    "data_even":  "FFFFFFFF",  # White
    "data_odd":   "FFF5F5F5",  # Very light gray
    "sai_col":    "FFEBF5E8",  # Light green for SAI column
    "source_col": "FFE8F4F8",  # Pale blue for source filename column
    "best_payoff": "FFFFEEBA", # Amber highlight for best (least negative) payoff per player
}

THIN = Side(style="thin", color="000000")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ---------------------------------------------------------------------------
# GDX parsing
# ---------------------------------------------------------------------------

class SymbolNotFoundError(RuntimeError):
    """Raised when gdxdump reports a symbol does not exist in the GDX file."""


def _gdxdump(gdx_path: Path, symbol: str) -> str:
    """
    Run gdxdump for one symbol and return stdout.

    Raises SymbolNotFoundError if gdxdump exits non-zero with no stderr output
    (which is how gdxdump signals an unknown symbol).
    Raises RuntimeError for any other failure.
    """
    result = subprocess.run(
        [str(GDXDUMP), str(gdx_path), f"Symb={symbol}"],
        capture_output=True, text=True,
    )
    if result.returncode != 0:
        stderr = result.stderr.strip()
        if not stderr:
            raise SymbolNotFoundError(
                f"Symbol '{symbol}' not found in {gdx_path.name}"
            )
        raise RuntimeError(
            f"gdxdump failed for {gdx_path.name} (symbol={symbol}):\n{stderr}"
        )
    return result.stdout


def parse_gdx_parameter(gdx_path: Path, symbol: str) -> dict[str, float]:
    """
    Parse a GAMS *parameter* from a GDX file.

    Lines look like:  'key' numeric_value,
    Returns dict mapping key (lowercase) → float value.
    """
    text = _gdxdump(gdx_path, symbol)
    pattern = re.compile(r"'([^']+)'\s+([-+\d.eE]+)")
    values: dict[str, float] = {}
    for line in text.splitlines():
        m = pattern.search(line)
        if m:
            values[m.group(1).strip().lower()] = float(m.group(2))
    if not values:
        raise RuntimeError(
            f"Symbol '{symbol}' not found or has no values in {gdx_path.name}"
        )
    return values


def parse_gdx_parameter_2d(
    gdx_path: Path, symbol: str
) -> dict[tuple[str, str], float]:
    """
    Parse a 2-dimensional GAMS *parameter* from a GDX file.

    Lines look like:  'dim1'.'dim2' numeric_value,
    Returns dict mapping (dim1_lower, dim2_lower) → float value.
    """
    text = _gdxdump(gdx_path, symbol)
    pattern = re.compile(r"'([^']+)'\.'([^']+)'\s+([-+\d.eE]+)")
    values: dict[tuple[str, str], float] = {}
    for line in text.splitlines():
        m = pattern.search(line)
        if m:
            values[(m.group(1).strip().lower(), m.group(2).strip().lower())] = float(
                m.group(3)
            )
    if not values:
        raise RuntimeError(
            f"Symbol '{symbol}' not found or has no values in {gdx_path.name}"
        )
    return values


def parse_gdx_variable_levels(gdx_path: Path, symbol: str) -> dict[str, float]:
    """
    Parse the level values (.L) of a GAMS *variable* from a GDX file.

    Lines look like:  'period'.L numeric_value,
    Periods with no .L entry are treated as zero (variable at default lower bound).
    Returns dict mapping period key (lowercase) → float level value.
    """
    text = _gdxdump(gdx_path, symbol)
    pattern = re.compile(r"'([^']+)'\.L\s+([-+\d.eE]+)")
    values: dict[str, float] = {}
    for line in text.splitlines():
        m = pattern.search(line)
        if m:
            values[m.group(1).strip().lower()] = float(m.group(2))
    # Missing periods → level is 0 (GAMS default)
    return values


def _in_year_range(
    calendar_year: float,
    start_year: int | None,
    end_year: int,
    *,
    end_inclusive: bool = True,
) -> bool:
    """Return True if year is in [start_year, end_year] or [start_year, end_year)."""
    lower_ok = start_year is None or calendar_year >= start_year
    upper_ok = calendar_year <= end_year if end_inclusive else calendar_year < end_year
    return lower_ok and upper_ok


def _sai_col_name(start_year: int | None, end_year: int, *, average: bool = False) -> str:
    """Return the W_SAI column label for the given year range."""
    agg = "avg" if average else "sum"
    if start_year is None:
        return f"W_SAI_{agg}_≤{end_year}"
    return f"W_SAI_{agg}_{start_year}-{end_year}"


def compute_sai_sum(
    gdx_path: Path,
    start_year: int | None,
    end_year: int,
    *,
    end_inclusive: bool = True,
    average: bool = False,
) -> float:
    """
    Aggregate W_SAI level values for all periods whose calendar year falls within
    [start_year, end_year] by default, or [start_year, end_year) when
    end_inclusive=False. start_year=None means no lower bound.

    When average=True, returns the mean across matched periods instead of the sum.
    Uses the `year` parameter from the same GDX file to map period indices to
    calendar years.  If W_SAI is absent from the GDX file (e.g. a no-deployment
    scenario), returns 0.0.
    """
    year_map = parse_gdx_parameter(gdx_path, YEAR_SYMBOL)
    try:
        sai_levels = parse_gdx_variable_levels(gdx_path, SAI_SYMBOL)
    except SymbolNotFoundError:
        return 0.0

    total = 0.0
    count = 0
    for period_key, calendar_year in year_map.items():
        if _in_year_range(
            calendar_year,
            start_year,
            end_year,
            end_inclusive=end_inclusive,
        ):
            total += sai_levels.get(period_key, 0.0)
            count += 1
    if average:
        if count == 0:
            return 0.0
        return total / count
    return total


def compute_welfare_sums(
    gdx_path: Path,
    region_codes: list[str],
    start_year: int | None,
    end_year: int,
    *,
    end_inclusive: bool = True,
    average: bool = False,
) -> dict[str, float]:
    """
    Aggregate `welfare_regional_t` for each region over all periods whose calendar
    year falls within [start_year, end_year] by default, or [start_year, end_year)
    when end_inclusive=False. start_year=None means no lower bound.

    When average=True, returns the mean across matched periods instead of the sum.

    Args:
        gdx_path:     Path to the GDX file.
        region_codes: List of lowercase GAMS region codes (e.g. ['nde', 'usa', 'chn']).
        start_year:   First year to include (None = no lower bound).
        end_year:     Last year bound.
        average:      If True, divide by the number of matched periods.

    Returns:
        Dict mapping region code (lowercase) → aggregated welfare float.
    """
    year_map = parse_gdx_parameter(gdx_path, YEAR_SYMBOL)
    welfare_by_region_period = parse_gdx_parameter_2d(gdx_path, WELFARE_SYMBOL)

    totals: dict[str, float] = {code.lower(): 0.0 for code in region_codes}
    count = 0
    for period_key, calendar_year in year_map.items():
        if _in_year_range(
            calendar_year,
            start_year,
            end_year,
            end_inclusive=end_inclusive,
        ):
            count += 1
            for code in region_codes:
                key = (period_key.lower(), code.lower())
                totals[code.lower()] += welfare_by_region_period.get(key, 0.0)
    if average and count > 0:
        for code in region_codes:
            totals[code.lower()] /= count
    return totals


def _build_global_token_map() -> dict[str, str]:
    """
    Build a mapping from all known RICE50x filename tokens (lowercase) to UPPERCASE
    display names, using the full RICE50X_REGIONS registry.

    Display names are the uppercase GDX codes (e.g. 'chn' → 'CHN').
    """
    from lib.rice50x_regions import RICE50X_REGIONS
    token_map: dict[str, str] = {}
    for gdx_code in RICE50X_REGIONS:
        display_name = gdx_code.upper()
        token_map[gdx_code.lower()] = display_name
    return token_map


# Module-level token map built once from all known RICE50x regions
_GLOBAL_TOKEN_MAP: dict[str, str] = _build_global_token_map()


def _parse_deployers(raw_token: str, token_map: dict[str, str]) -> set[str]:
    """
    Greedily parse a raw deployer token (e.g. 'usarusind') into a set of display
    names (e.g. {'USA', 'RUS', 'IND'}).  Longest tokens are matched first to avoid
    accidental partial matches.

    Raises ValueError if the token cannot be fully consumed.
    """
    tokens_by_length = sorted(token_map.keys(), key=len, reverse=True)
    remaining = raw_token.lower()
    deployers: set[str] = set()
    while remaining:
        for token in tokens_by_length:
            if remaining.startswith(token):
                deployers.add(token_map[token])
                remaining = remaining[len(token):]
                break
        else:
            raise ValueError(
                f"Cannot parse deployer token '{raw_token}': "
                f"unrecognised segment '{remaining}'"
            )
    return deployers


def _deployers_to_key(deployers: set[str]) -> str:
    """
    Convert a set of deploying display names to a payoff table row key.

    Returns '( )' for the empty set (no deployment), or '(MEMBERS)' with
    members sorted alphabetically for any non-empty set (including singletons).

    Examples:
        set()           → '( )'
        {'IND'}         → '(IND)'
        {'IND', 'RUS'}  → '(INDRUS)'
        {'IND','RUS','USA'} → '(INDRUSUSA)'
    """
    if not deployers:
        return "( )"
    return "(" + "".join(sorted(deployers)) + ")"


def _all_deployer_keys(players: list[str]) -> list[str]:
    """Return all 2^n deployer subset keys in order (empty first, then by size)."""
    result = []
    for r in range(len(players) + 1):
        for subset in combinations(sorted(players), r):
            result.append(_deployers_to_key(set(subset)))
    return result


def _detect_common_prefix(stems: list[str]) -> str:
    """Return the longest common prefix shared by all stems."""
    if not stems:
        return ""
    prefix = stems[0]
    for stem in stems[1:]:
        # Shorten prefix until it matches
        while not stem.startswith(prefix):
            prefix = prefix[:-1]
            if not prefix:
                return ""
    return prefix


def discover_state_files(
    input_dir: Path,
    required_stem_prefix: str | None = None,
) -> tuple[dict[str, Path], list[tuple[Path, str]]]:
    """
    Scan GDX files in input_dir and map each to a deployer key.

    Uses the full RICE50X_REGIONS registry to parse any known country combination
    from filenames, regardless of the players list passed to ingest().

    Strategy:
      1. Strip '_deployed' / '_deploy' suffix from each stem.
      2. Auto-detect the longest common prefix of all stripped stems.
      3. Each file's deployer token = stripped_stem[len(prefix):].lstrip('_').
      4. Empty token → '( )' (no-deployment state).
      5. Parse token into a set of display names via greedy matching.
      6. Any deployer set (including singletons) → '(MEMBERS)' key.

    Args:
        input_dir: Directory containing GDX files.
        required_stem_prefix: Optional filename-stem prefix filter (e.g.
            'results_ssp2_bau_impact_andreoni'). If provided, only files whose
            stem starts with this prefix are considered.

    Returns:
        key_to_file : dict mapping deployer key → GDX Path
        skipped     : list of (Path, reason) for files that could not be parsed
    """
    gdx_files = sorted(input_dir.glob("*.gdx"))
    if not gdx_files:
        raise FileNotFoundError(f"No .gdx files found in {input_dir}")

    if required_stem_prefix is not None:
        gdx_files = [p for p in gdx_files if p.stem.startswith(required_stem_prefix)]
        if not gdx_files:
            raise FileNotFoundError(
                f"No .gdx files with stem prefix '{required_stem_prefix}' found in {input_dir}"
            )

    # Prefer the main results_* files if present; ignore debug_iter_* helpers.
    if any(p.stem.startswith("results_") for p in gdx_files):
        gdx_files = [p for p in gdx_files if p.stem.startswith("results_")]

    token_map = _GLOBAL_TOKEN_MAP

    # Strip deployment suffix from stems
    deploy_suffixes = ["_deployed", "_deploy"]
    stripped: dict[Path, str] = {}
    for gp in gdx_files:
        stem = gp.stem
        for sfx in deploy_suffixes:
            if stem.endswith(sfx):
                stem = stem[: -len(sfx)]
                break
        stripped[gp] = stem

    # Auto-detect common prefix
    common_prefix = _detect_common_prefix(list(stripped.values()))

    key_to_file: dict[str, Path] = {}
    skipped: list[tuple[Path, str]] = []
    for gp, stem in stripped.items():
        deployer_token = stem[len(common_prefix):].lstrip("_")
        if not deployer_token:
            key = "( )"
        else:
            try:
                deployers = _parse_deployers(deployer_token, token_map)
            except ValueError as exc:
                skipped.append((gp, str(exc)))
                continue
            key = _deployers_to_key(deployers)

        if key in key_to_file:
            raise RuntimeError(
                f"Two GDX files map to the same key '{key}': "
                f"{key_to_file[key].name} and {gp.name}"
            )
        key_to_file[key] = gp

    return key_to_file, skipped


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

def write_payoff_table(
    df: pd.DataFrame,
    output_path: Path,
    players: list[tuple[str, str]],
    source_dir: str,
    start_year: int | None,
    end_year: int,
    extra_metadata: dict[str, object] | None = None,
    average: bool = False,
) -> None:
    """
    Write the payoff DataFrame to a formatted Excel file.

    Args:
        df: DataFrame indexed by state name, columns:
              <display_name>...  (welfare per player)
              W_SAI_sum          (summed SAI deployment)
              source_file        (original filename)
        output_path: Destination .xlsx path.
        players: List of (gdx_code, display_name) tuples.
        source_dir: Original input directory (written to metadata sheet).
        start_year: First year included (None = no lower bound).
        end_year:   Last year included (inclusive).
        extra_metadata: Optional additional metadata key/value pairs.
    """
    display_names = [d for _, d in players]
    sai_col = _sai_col_name(start_year, end_year, average=average)
    states = df.index.tolist()
    n_players = len(display_names)

    # Column order in the sheet
    data_cols = display_names + [sai_col]
    total_cols = 1 + len(data_cols) + 1  # state + data + source

    wb = Workbook()
    ws = wb.active
    ws.title = "Payoffs"

    # Column widths
    ws.column_dimensions["A"].width = 16
    for i in range(n_players):
        ws.column_dimensions[get_column_letter(2 + i)].width = 18
    # SAI column
    ws.column_dimensions[get_column_letter(2 + n_players)].width = 20
    # Source column
    ws.column_dimensions[get_column_letter(3 + n_players)].width = 42

    # ------------------------------------------------------------------
    # Row 1: Title (merged)
    # ------------------------------------------------------------------
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    agg_label = "avg (per period)" if sai_col.startswith("W_SAI_avg") else "sum"
    title_cell = ws.cell(
        row=1, column=1,
        value=f"Precomputed Payoffs [{agg_label}] — {WELFARE_SYMBOL}  |  {SAI_SYMBOL}: {sai_col}",
    )
    title_cell.font = Font(name="Calibri", bold=True, size=12, color="FFFFFFFF")
    title_cell.fill = PatternFill(start_color=COLORS["title"],
                                  end_color=COLORS["title"], fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    # ------------------------------------------------------------------
    # Row 2: Column headers
    # ------------------------------------------------------------------
    headers = ["State"] + data_cols + ["Source file"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(name="Calibri", bold=True, size=10)
        # SAI column gets its own header color
        if header == sai_col:
            cell.fill = PatternFill(start_color=COLORS["sai_col"],
                                    end_color=COLORS["sai_col"], fill_type="solid")
        else:
            cell.fill = PatternFill(start_color=COLORS["header"],
                                    end_color=COLORS["header"], fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER_ALL
    ws.row_dimensions[2].height = 16

    # ------------------------------------------------------------------
    # Data rows
    # ------------------------------------------------------------------
    for row_offset, state in enumerate(states):
        row_idx = 3 + row_offset
        row_color = COLORS["data_even"] if row_offset % 2 == 0 else COLORS["data_odd"]

        # State name
        state_cell = ws.cell(row=row_idx, column=1, value=state)
        state_cell.font = Font(name="Calibri", bold=True, size=10)
        state_cell.fill = PatternFill(start_color=COLORS["state_name"],
                                      end_color=COLORS["state_name"], fill_type="solid")
        state_cell.alignment = Alignment(horizontal="center", vertical="center")
        state_cell.border = BORDER_ALL

        # Welfare columns (one per player)
        for p_idx, player in enumerate(display_names):
            cell = ws.cell(row=row_idx, column=2 + p_idx, value=df.loc[state, player])
            cell.number_format = "0.000000"
            cell.font = Font(name="Calibri", size=10)
            cell.fill = PatternFill(start_color=row_color,
                                    end_color=row_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = BORDER_ALL

        # SAI sum column
        sai_cell = ws.cell(row=row_idx, column=2 + n_players,
                           value=df.loc[state, sai_col])
        sai_cell.number_format = "0.000"
        sai_cell.font = Font(name="Calibri", size=10)
        sai_cell.fill = PatternFill(start_color=COLORS["sai_col"],
                                    end_color=COLORS["sai_col"], fill_type="solid")
        sai_cell.alignment = Alignment(horizontal="right", vertical="center")
        sai_cell.border = BORDER_ALL

        # Source file
        src_cell = ws.cell(row=row_idx, column=3 + n_players,
                           value=df.loc[state, "source_file"])
        src_cell.font = Font(name="Calibri", size=9, italic=True, color="FF555555")
        src_cell.fill = PatternFill(start_color=COLORS["source_col"],
                                    end_color=COLORS["source_col"], fill_type="solid")
        src_cell.alignment = Alignment(horizontal="left", vertical="center")
        src_cell.border = BORDER_ALL

        ws.row_dimensions[row_idx].height = 15

    # ------------------------------------------------------------------
    # Highlight top-3 payoffs (least negative) per player column
    # ------------------------------------------------------------------
    rank_fills = [
        PatternFill(start_color="FFFFD966", end_color="FFFFD966", fill_type="solid"),  # 1st: gold
        PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid"),  # 2nd: light gold
        PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid"),  # 3rd: pale gold
    ]
    for p_idx, player in enumerate(display_names):
        col_idx = 2 + p_idx
        sorted_vals = sorted(df[player].dropna().unique(), reverse=True)
        top_vals = sorted_vals[:3]
        for row_offset, state in enumerate(states):
            val = df.loc[state, player]
            for rank, top_val in enumerate(top_vals):
                if math.isclose(val, float(top_val), rel_tol=1e-13):
                    ws.cell(row=3 + row_offset, column=col_idx).fill = rank_fills[rank]
                    break

    # ------------------------------------------------------------------
    # Metadata sheet
    # ------------------------------------------------------------------
    ws_meta = wb.create_sheet(title="Metadata")
    ws_meta.column_dimensions["A"].width = 25
    ws_meta.column_dimensions["B"].width = 50

    meta_rows = [
        ("Parameter", "Value"),
        ("welfare_symbol", WELFARE_SYMBOL),
        ("sai_symbol", SAI_SYMBOL),
        ("sai_start_year", start_year if start_year is not None else ""),
        ("sai_end_year", end_year),
        ("source_dir", source_dir),
        ("num_states", len(states)),
        ("num_players", n_players),
        ("players", ", ".join(display_names)),
        ("gdx_region_codes", ", ".join(code for code, _ in players)),
        ("states", ", ".join(states)),
    ]
    if extra_metadata:
        for key, value in extra_metadata.items():
            meta_rows.append((str(key), "" if value is None else value))

    for r_idx, (key, value) in enumerate(meta_rows, start=1):
        is_header = r_idx == 1
        for c_idx, text in enumerate([key, value], start=1):
            cell = ws_meta.cell(row=r_idx, column=c_idx, value=str(text))
            cell.font = Font(name="Calibri", bold=is_header, size=10)
            if is_header:
                cell.fill = PatternFill(start_color=COLORS["header"],
                                        end_color=COLORS["header"], fill_type="solid")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = BORDER_ALL

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))


# ---------------------------------------------------------------------------
# Main logic
# ---------------------------------------------------------------------------

def _check_sai_ordering(df: "pd.DataFrame", sai_col: str) -> None:
    """
    Warn if a coalition's W_SAI sum falls outside the range of its members' singletons.

    For each multi-player coalition state present in df, find the singleton W_SAI
    values for each member and check that the coalition value lies between the
    minimum and maximum singleton value.  Missing singletons are skipped (no warning).
    """
    warnings: list[str] = []
    for key in df.index:
        if key == "( )":
            continue
        inner = key[1:-1]  # strip parentheses, e.g. 'CHNNDE'
        try:
            members = _parse_deployers(inner, _GLOBAL_TOKEN_MAP)
        except ValueError:
            continue
        if len(members) < 2:
            continue

        singleton_keys = [_deployers_to_key({m}) for m in members]
        available = [k for k in singleton_keys if k in df.index]
        if len(available) < 2:
            continue  # not enough singletons to bound

        singleton_sai = [df.loc[k, sai_col] for k in available]
        coalition_sai = df.loc[key, sai_col]
        lo, hi = min(singleton_sai), max(singleton_sai)

        if not (lo <= coalition_sai <= hi):
            warnings.append(
                f"  {key}: W_SAI={coalition_sai:.2f} is outside [{lo:.2f}, {hi:.2f}] "
                f"(members: {', '.join(sorted(members))})"
            )

    if warnings:
        print("\nWarning: unexpected W_SAI ordering for the following coalitions:")
        for w in warnings:
            print(w)
        print("  A coalition's SAI should lie between the min and max of its members' singletons.")
    else:
        print("\nW_SAI ordering check passed.")


def ingest(
    input_dir: Path,
    output_path: Path,
    players: list[tuple[str, str]] | None,
    start_year: int | None,
    end_year: int,
    end_inclusive: bool = True,
    extra_metadata: dict[str, object] | None = None,
    required_stem_prefix: str | None = None,
    average: bool = False,
) -> None:
    sai_col = _sai_col_name(start_year, end_year, average=average)

    # --- Discover which GDX files map to which deployer keys ------------------
    key_to_file, skipped = discover_state_files(
        input_dir,
        required_stem_prefix=required_stem_prefix,
    )

    # --- Auto-detect players from discovered deployers if not provided --------
    if players is None:
        all_names: set[str] = set()
        for key in key_to_file:
            if key != "( )":
                # key is like '(CHN)', '(CHNNDE)', '(CHNNDEUSA)' etc.
                inner = key[1:-1]  # strip parentheses
                # Re-parse inner string using global token map to get individual names
                all_names.update(_parse_deployers(inner, _GLOBAL_TOKEN_MAP))
        if not all_names:
            raise ValueError(
                f"Could not auto-detect players from GDX files in {input_dir}. "
                "No deployer states found."
            )
        # GDX code = lowercase of display name (RICE50x convention)
        players = sorted([(name.lower(), name) for name in all_names], key=lambda x: x[1])
        print(f"Auto-detected players: {[name for _, name in players]}")

    # Normalise players list so display names are always uppercase
    players = [(code, name.upper()) for code, name in players]
    display_names = [name for _, name in players]

    total_gdx = len(list(input_dir.glob("*.gdx")))
    print(f"Found {total_gdx} GDX file(s) in {input_dir}")
    end_bracket = "]" if end_inclusive else ")"
    if start_year is None:
        print(f"Summing {SAI_SYMBOL} for years (-∞, {end_year}{end_bracket}\n")
    else:
        print(f"Summing {SAI_SYMBOL} for years [{start_year}, {end_year}{end_bracket}\n")

    if skipped:
        print("Skipped files:")
        for path, reason in skipped:
            print(f"  {path.name}: {reason}")
        print()

    # --- Check all 2^n deployer subsets are covered ---------------------------
    all_keys = _all_deployer_keys(display_names)
    missing = [k for k in all_keys if k not in key_to_file]
    if missing:
        print(
            f"Warning: the following deployer subsets have no GDX file and will be "
            f"omitted from the table: {missing}"
        )

    # Process rows in subset order (empty first, then singletons, pairs, grand)
    ordered_keys = [k for k in all_keys if k in key_to_file]

    region_codes = [code for code, _ in players]

    # --- Extract payoffs from each GDX file -----------------------------------
    rows: list[dict] = []
    for state_name in ordered_keys:
        gdx_path = key_to_file[state_name]
        print(f"  {gdx_path.name}  →  deployer '{state_name}'")

        welfare = compute_welfare_sums(
            gdx_path,
            region_codes,
            start_year,
            end_year,
            end_inclusive=end_inclusive,
            average=average,
        )
        sai_sum = compute_sai_sum(
            gdx_path,
            start_year,
            end_year,
            end_inclusive=end_inclusive,
            average=average,
        )

        row: dict = {"state": state_name, "source_file": gdx_path.name}
        for code, display in players:
            if code.lower() not in welfare:
                raise KeyError(
                    f"Region '{code}' not found in {gdx_path.name}. "
                    f"Available keys: {sorted(welfare.keys())}"
                )
            row[display] = welfare[code.lower()]
        row[sai_col] = sai_sum
        rows.append(row)

    df = pd.DataFrame(rows).set_index("state")
    df = df[display_names + [sai_col, "source_file"]]

    _check_sai_ordering(df, sai_col)

    combined_metadata: dict[str, object] = {"welfare_aggregation": "average" if average else "sum"}
    if extra_metadata:
        combined_metadata.update(extra_metadata)

    write_payoff_table(
        df,
        output_path,
        players,
        str(input_dir),
        start_year,
        end_year,
        extra_metadata=combined_metadata,
        average=average,
    )

    print(f"\nWrote payoff table → {output_path}")
    print(df[display_names + [sai_col]].to_string())


def main() -> None:
    repo_root = Path(__file__).parent.parent

    parser = argparse.ArgumentParser(
        description=(
            "Ingest welfare_regional_farsighted and W_SAI from GDX files "
            "into a payoff table Excel file."
        )
    )
    parser.add_argument(
        "--input-dir",
        default=str(DEFAULT_INPUT_DIR),
        help="Directory containing .gdx files (default: %(default)s)",
    )
    parser.add_argument(
        "--output",
        default="burke",
        help=(
            "Output path or bare stem. A bare name (no directory) is placed in "
            "payoff_tables/; .xlsx is appended if missing. (default: %(default)s)"
        ),
    )
    parser.add_argument(
        "--year-range",
        default=str(DEFAULT_CUTOFF_YEAR),
        help=(
            "Year range for summation. Either a single end year (e.g. '2060', "
            "meaning all periods up to and including 2060) or a start-end range "
            "(e.g. '2060-2080', inclusive on both ends). (default: %(default)s)"
        ),
    )
    parser.add_argument(
        "--players",
        default=None,
        help=(
            "Comma-separated gdx_code:display_name pairs, "
            "e.g. 'nde:IND,usa:USA,rus:RUS' (default: nde:IND,usa:USA,rus:RUS)"
        ),
    )
    args = parser.parse_args()

    # Parse --year-range: either 'YYYY' or 'YYYY-YYYY'
    year_range = args.year_range.strip()
    if "-" in year_range:
        parts = year_range.split("-", 1)
        start_year: int | None = int(parts[0])
        end_year = int(parts[1])
    else:
        start_year = None
        end_year = int(year_range)

    if args.players:
        players = []
        for part in args.players.split(","):
            code, display = part.strip().split(":")
            players.append((code.strip(), display.strip()))
    else:
        players = DEFAULT_PLAYERS

    output = Path(args.output)
    if not output.suffix:
        output = output.with_suffix(".xlsx")
    if not output.parent.parts or output.parent == Path("."):
        output = repo_root / "payoff_tables" / output

    ingest(
        input_dir=Path(args.input_dir),
        output_path=output,
        players=players,
        start_year=start_year,
        end_year=end_year,
    )


if __name__ == "__main__":
    main()
