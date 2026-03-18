"""
Custom Excel writer for strategy tables that exactly matches the original format.

This creates Excel files from scratch to match the exact structure, styling,
and formatting of the original strategy tables.
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import math
import pandas as pd
import numpy as np
import hashlib
import json
from pathlib import Path


# Color scheme (from original files)
COLORS = {
    'proposer_header': 'FF4D9CC9',      # Blue for "Proposer X"
    'state_name': 'FFEEBF99',           # Beige/orange for state names in column A
    'proposition_row': 'FF99C7E0',      # Light blue for proposition rows
    'acceptance_row': 'FFCCECE3',       # Very light blue/green for acceptance rows
    'player_name': 'FF66C5AB',          # Green for player names in acceptance rows
    'self_acceptance': 'FFE28E4D',      # Orange for player accepting own proposal
    'nan_cell': 'FFD9D9D9',             # Gray for NaN/None cells
}


def generate_config_hash(config, length=6):
    """
    Generate a short hash from configuration parameters.

    Args:
        config: Configuration dictionary
        length: Length of hash to return (default: 6)

    Returns:
        Short hash string (e.g., 'a3f2b1')
    """
    # Create a sorted JSON string of relevant config parameters.
    # Exclude metadata fields that don't affect the game (scenario_name,
    # scenario_description, state_names).
    hash_params = {k: v for k, v in sorted(config.items())
                   if k not in ['scenario_name', 'state_names', 'scenario_description']}

    # Convert to JSON string (with sorted keys for consistency)
    config_str = json.dumps(hash_params, sort_keys=True, default=str)

    # Generate hash
    hash_obj = hashlib.md5(config_str.encode())
    return hash_obj.hexdigest()[:length]


def generate_filename(config, description=None, output_dir='./strategy_tables'):
    """
    Generate filename from configuration parameters.

    Format: eq_n{n}_{scenario_name_without_nN_suffix}_{hash}.xlsx
    Example: eq_n4_power_threshold_a3f2b1.xlsx

    Args:
        config: Configuration dictionary
        description: Unused (kept for backwards compatibility)
        output_dir: Output directory

    Returns:
        Full path to output file
    """
    # Number of players
    n = len(config['players'])

    # Scenario name: strip the trailing _nN suffix since n is already in the prefix
    import re
    scenario_part = re.sub(r'_n\d+$', '', config.get('scenario_name', 'unknown'))

    # Generate hash
    config_hash = generate_config_hash(config)

    # Build filename parts
    parts = [
        'eq',
        f'n{n}',
        scenario_part,
        config_hash,
    ]

    # Combine into filename
    filename = '_'.join(parts) + '.xlsx'

    return str(Path(output_dir) / filename)


def write_strategy_table_excel(df: pd.DataFrame, excel_file_path: str, players: list,
                              effectivity: dict = None, states: list = None, metadata: dict = None,
                              value_functions: pd.DataFrame = None, geo_levels: dict = None,
                              deploying_coalitions: dict = None, static_payoffs: pd.DataFrame = None,
                              transition_matrix: pd.DataFrame = None):
    """
    Write a strategy DataFrame to Excel with exact formatting matching original files.

    Args:
        df: Strategy DataFrame with MultiIndex structure:
            - Index: (Current State, Type, Player)
            - Columns: (Proposer, Next State)
        excel_file_path: Output file path
        players: List of player names (e.g., ['W', 'T', 'C'])
        effectivity: Effectivity correspondence dict (optional)
        states: List of state names (optional, inferred from df if not provided)
        metadata: Dictionary of configuration parameters to save in metadata sheet (optional)
        value_functions: DataFrame with states as index and players as columns (optional)
        geo_levels: Dictionary mapping state names to geoengineering levels (optional)
        deploying_coalitions: Dictionary mapping state names to deploying coalition names (optional)
    """
    # Create new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Strategy"

    # Extract structure from DataFrame
    if states is None:
        states = df.index.get_level_values(0).unique().tolist()
    proposers = df.columns.get_level_values(0).unique().tolist()

    # Helper function to check if transition is unilateral
    def is_unilateral_transition(current_state, next_state, proposer_label, effectivity_dict):
        """
        Check if a transition can be done without consulting others.

        Unilateral transitions:
        1. Maintaining status quo (current_state == next_state)
        2. Leaving a coalition (if only proposer or no one needs to approve)
        """
        # Extract proposer name from label (e.g., "Proposer W" -> "W")
        proposer_name = proposer_label.split()[-1] if 'Proposer' in proposer_label else proposer_label

        # Same state is always unilateral (maintaining status quo)
        if current_state == next_state:
            return True

        # If we don't have effectivity info, can't determine
        if effectivity_dict is None:
            return False

        # Check approval committee for this transition
        # Effectivity key: (proposer_name, current_state, next_state, responder)
        # If responder in committee, effectivity[(proposer_name, current_state, next_state, responder)] == 1
        approval_committee = []
        for player in players:
            key = (proposer_name, current_state, next_state, player)
            if key in effectivity_dict and effectivity_dict[key] == 1:
                approval_committee.append(player)

        # Unilateral if:
        # - Empty approval committee, OR
        # - Only proposer in approval committee
        is_unilateral = (len(approval_committee) == 0 or
                        (len(approval_committee) == 1 and proposer_name in approval_committee))

        return is_unilateral

    # Set column widths
    ws.column_dimensions['A'].width = 14.0
    ws.column_dimensions['B'].width = 8.77734375
    ws.column_dimensions['C'].width = 3.77734375

    # Data columns: uniform width of 7.0 (approximately 1.85 cm)
    for col_idx in range(4, 4 + len(proposers) * len(states)):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 7.0

    # Set row heights
    ws.row_dimensions[1].height = 16.5
    ws.row_dimensions[2].height = 14.85
    ws.row_dimensions[3].height = 14.85

    # Write headers
    # Row 1: Proposer headers (merged cells)
    col = 4
    for proposer_label in proposers:
        # Merge cells for this proposer (spans len(states) columns)
        start_col = col
        end_col = col + len(states) - 1
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

        # Write proposer name in merged cell
        ws.cell(row=1, column=start_col, value=proposer_label)
        ws.cell(row=1, column=start_col).fill = PatternFill(
            start_color=COLORS['proposer_header'],
            end_color=COLORS['proposer_header'],
            fill_type='solid'
        )
        ws.cell(row=1, column=start_col).font = Font(name='Calibri', bold=True, size=9)
        ws.cell(row=1, column=start_col).alignment = Alignment(horizontal='center', vertical='center')

        col = end_col + 1

    # Row 2: State headers (repeated for each proposer)
    col = 4
    for proposer_label in proposers:
        for state in states:
            ws.cell(row=2, column=col, value=state)
            ws.cell(row=2, column=col).font = Font(name='Calibri', size=9)
            ws.cell(row=2, column=col).alignment = Alignment(horizontal='center', vertical='center')
            col += 1

    # Write data rows
    current_row = 3
    for state in states:
        # Proposition row
        ws.cell(row=current_row, column=1, value=state)
        ws.cell(row=current_row, column=1).fill = PatternFill(
            start_color=COLORS['state_name'],
            end_color=COLORS['state_name'],
            fill_type='solid'
        )
        ws.cell(row=current_row, column=1).font = Font(name='Calibri', bold=True, size=9)
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='right', vertical='center')

        ws.cell(row=current_row, column=2, value='Proposition')
        ws.cell(row=current_row, column=2).fill = PatternFill(
            start_color=COLORS['proposition_row'],
            end_color=COLORS['proposition_row'],
            fill_type='solid'
        )
        ws.cell(row=current_row, column=2).font = Font(name='Calibri', size=9)
        ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center', vertical='center')

        # Column 3: Write string 'NaN' to prevent pandas forward-fill when reading
        ws.cell(row=current_row, column=3, value='NaN')
        ws.cell(row=current_row, column=3).fill = PatternFill(
            start_color=COLORS['proposition_row'],
            end_color=COLORS['proposition_row'],
            fill_type='solid'
        )
        ws.cell(row=current_row, column=3).font = Font(name='Calibri', size=9)
        ws.cell(row=current_row, column=3).alignment = Alignment(horizontal='center', vertical='center')

        # Fill proposition probabilities
        col = 4
        for proposer_label in proposers:
            for next_state in states:
                val = df.loc[(state, 'Proposition', np.nan), (proposer_label, next_state)]
                # Only show values != 0
                if pd.isna(val) or val == 0:
                    ws.cell(row=current_row, column=col, value=None)
                else:
                    ws.cell(row=current_row, column=col, value=float(val))
                    ws.cell(row=current_row, column=col).number_format = '0.###'

                ws.cell(row=current_row, column=col).fill = PatternFill(
                    start_color=COLORS['proposition_row'],
                    end_color=COLORS['proposition_row'],
                    fill_type='solid'
                )
                ws.cell(row=current_row, column=col).font = Font(name='Calibri', size=9)
                ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='center', vertical='center')
                col += 1

        current_row += 1

        # Acceptance rows (one per player)
        for player in players:
            # State name (all acceptance rows)
            ws.cell(row=current_row, column=1, value=state)
            ws.cell(row=current_row, column=1).fill = PatternFill(
                start_color=COLORS['state_name'],
                end_color=COLORS['state_name'],
                fill_type='solid'
            )
            ws.cell(row=current_row, column=1).font = Font(name='Calibri', bold=True, size=9)
            ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='right', vertical='center')

            # Type (all acceptance rows)
            ws.cell(row=current_row, column=2, value='Acceptance')
            ws.cell(row=current_row, column=2).fill = PatternFill(
                start_color=COLORS['acceptance_row'],
                end_color=COLORS['acceptance_row'],
                fill_type='solid'
            )
            ws.cell(row=current_row, column=2).font = Font(name='Calibri', size=9)
            ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center', vertical='center')

            # Player name
            ws.cell(row=current_row, column=3, value=player)
            ws.cell(row=current_row, column=3).fill = PatternFill(
                start_color=COLORS['player_name'],
                end_color=COLORS['player_name'],
                fill_type='solid'
            )
            ws.cell(row=current_row, column=3).font = Font(name='Calibri', size=9)
            ws.cell(row=current_row, column=3).alignment = Alignment(horizontal='center', vertical='center')

            # Fill acceptance probabilities
            col = 4
            for proposer_label in proposers:
                proposer_name = proposer_label.split()[-1]  # Extract "W" from "Proposer W"

                for next_state in states:
                    val = df.loc[(state, 'Acceptance', player), (proposer_label, next_state)]

                    # Determine whether this cell SHOULD be in the approval committee.
                    # When effectivity is provided, use the rule-based answer (the "ought").
                    # Fall back to NaN detection only when no effectivity dict is given.
                    if effectivity is not None:
                        eff_key = (proposer_name, state, next_state, player)
                        should_be_in_committee = effectivity.get(eff_key, 0) == 1
                    else:
                        should_be_in_committee = not pd.isna(val)

                    if not should_be_in_committee:
                        # Grey: Not in approval committee per rules
                        ws.cell(row=current_row, column=col, value=None)
                        ws.cell(row=current_row, column=col).fill = PatternFill(
                            start_color=COLORS['nan_cell'],
                            end_color=COLORS['nan_cell'],
                            fill_type='solid'
                        )
                    else:
                        # In approval committee per rules: write actual value (may be NaN
                        # if missing from file — cell will appear coloured but empty).
                        if not pd.isna(val):
                            ws.cell(row=current_row, column=col, value=float(val))
                            ws.cell(row=current_row, column=col).number_format = '0.###'
                        else:
                            ws.cell(row=current_row, column=col, value=None)

                        # Dark orange if: player == proposer AND unilateral transition
                        # (transitions the proposer can do without consulting others)
                        is_unilateral = is_unilateral_transition(state, next_state, proposer_label, effectivity)
                        is_self = (player == proposer_name)

                        if is_self and is_unilateral:
                            # Dark orange: Player accepting their own unilateral action
                            ws.cell(row=current_row, column=col).fill = PatternFill(
                                start_color=COLORS['self_acceptance'],  # FFE28E4D
                                end_color=COLORS['self_acceptance'],
                                fill_type='solid'
                            )
                        else:
                            # Light green: In approval committee per rules
                            ws.cell(row=current_row, column=col).fill = PatternFill(
                                start_color=COLORS['acceptance_row'],  # FFCCECE3
                                end_color=COLORS['acceptance_row'],
                                fill_type='solid'
                            )

                    ws.cell(row=current_row, column=col).font = Font(name='Calibri', size=9)
                    ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='center', vertical='center')

                    col += 1

            current_row += 1

    # Add borders
    # Define border styles
    thin_border = Side(style='thin', color='000000')

    # Determine table dimensions
    last_row = current_row - 1
    last_col = 3 + len(proposers) * len(states)  # 3 label cols + data cols

    # Calculate state group boundaries (horizontal borders between states)
    # Each state has 1 proposition row + len(players) acceptance rows
    rows_per_state = 1 + len(players)
    state_boundaries_rows = []
    for state_idx in range(len(states) - 1):  # Don't add border after last state
        boundary_row = 2 + (state_idx + 1) * rows_per_state  # 2 header rows + state rows
        state_boundaries_rows.append(boundary_row)

    # Apply borders
    for row in range(1, last_row + 1):
        for col in range(1, last_col + 1):
            cell = ws.cell(row=row, column=col)
            border_kwargs = {}

            # Outside border (medium)
            if row == 1:
                border_kwargs['top'] = thin_border
            if row == last_row:
                border_kwargs['bottom'] = thin_border
            if col == 1:
                border_kwargs['left'] = thin_border
            if col == last_col:
                border_kwargs['right'] = thin_border

            # Vertical border between label columns (A-C) and data columns (D onwards)
            if col == 3:
                border_kwargs['right'] = thin_border
            if col == 4:
                border_kwargs['left'] = thin_border

            # Vertical borders between proposers (after columns H, M, etc.)
            proposer_boundaries = [3 + len(states) * (i + 1) for i in range(len(proposers) - 1)]
            if col in proposer_boundaries:
                border_kwargs['right'] = thin_border
            if col - 1 in proposer_boundaries and col > 3:
                border_kwargs['left'] = thin_border

            # Horizontal border between header rows (1-2) and data rows (3+)
            if row == 2:
                border_kwargs['bottom'] = thin_border
            if row == 3:
                border_kwargs['top'] = thin_border

            # Horizontal borders between states
            if row in state_boundaries_rows:
                border_kwargs['bottom'] = thin_border
            if row - 1 in state_boundaries_rows:
                border_kwargs['top'] = thin_border

            # Apply border if any
            if border_kwargs:
                # Preserve existing borders (take the stronger one if both exist)
                existing_border = cell.border
                final_border = Border(
                    left=border_kwargs.get('left', existing_border.left),
                    right=border_kwargs.get('right', existing_border.right),
                    top=border_kwargs.get('top', existing_border.top),
                    bottom=border_kwargs.get('bottom', existing_border.bottom)
                )
                cell.border = final_border

    # Precompute geo_dict (used in both Long-term Values and Short-term Values sheets)
    geo_dict = None
    if geo_levels is not None:
        if isinstance(geo_levels, pd.DataFrame):
            geo_dict = geo_levels['G'].to_dict()
        else:
            geo_dict = geo_levels

    # Add Long-term Values sheet if value functions or geo levels are provided
    if value_functions is not None or geo_levels is not None:
        ws_results = wb.create_sheet(title="Long-term Values")

        # Set column widths
        ws_results.column_dimensions['A'].width = 12
        num_cols = 1 + len(players) + (1 if geo_dict else 0) + (1 if deploying_coalitions else 0)
        for i in range(1, num_cols):
            col_letter = get_column_letter(1 + i)
            ws_results.column_dimensions[col_letter].width = 12

        # Title
        ws_results.cell(row=1, column=1, value="Long-term Value Functions (V) and Geoengineering Levels")
        ws_results.cell(row=1, column=1).font = Font(name='Calibri', bold=True, size=12)
        ws_results.cell(row=1, column=1).fill = PatternFill(
            start_color='FF4D9CC9', end_color='FF4D9CC9', fill_type='solid'
        )
        # Merge title across all columns
        ws_results.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)

        # Headers
        current_row = 2
        ws_results.cell(row=current_row, column=1, value="State")
        ws_results.cell(row=current_row, column=1).font = Font(name='Calibri', bold=True, size=10)
        ws_results.cell(row=current_row, column=1).fill = PatternFill(
            start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid'
        )

        col_idx = 2
        # Player columns
        if value_functions is not None:
            for player in players:
                ws_results.cell(row=current_row, column=col_idx, value=player)
                ws_results.cell(row=current_row, column=col_idx).font = Font(name='Calibri', bold=True, size=10)
                ws_results.cell(row=current_row, column=col_idx).fill = PatternFill(
                    start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid'
                )
                ws_results.cell(row=current_row, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')
                col_idx += 1

        # G column
        if geo_dict is not None:
            ws_results.cell(row=current_row, column=col_idx, value="G (°C cooling)")
            ws_results.cell(row=current_row, column=col_idx).font = Font(name='Calibri', bold=True, size=10)
            ws_results.cell(row=current_row, column=col_idx).fill = PatternFill(
                start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid'
            )
            ws_results.cell(row=current_row, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')
            col_idx += 1

        # Deployed by column
        if deploying_coalitions is not None:
            ws_results.cell(row=current_row, column=col_idx, value="Deployed by")
            ws_results.cell(row=current_row, column=col_idx).font = Font(name='Calibri', bold=True, size=10)
            ws_results.cell(row=current_row, column=col_idx).fill = PatternFill(
                start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid'
            )
            ws_results.cell(row=current_row, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')

        # Data rows
        current_row = 3
        for state in states:
            col_idx = 1

            # State name
            ws_results.cell(row=current_row, column=col_idx, value=state)
            ws_results.cell(row=current_row, column=col_idx).font = Font(name='Calibri', size=10)
            ws_results.cell(row=current_row, column=col_idx).fill = PatternFill(
                start_color='FFEEBF99', end_color='FFEEBF99', fill_type='solid'
            )
            col_idx += 1

            # Value function columns
            if value_functions is not None:
                for player in players:
                    val = value_functions.loc[state, player]
                    ws_results.cell(row=current_row, column=col_idx, value=float(val))
                    ws_results.cell(row=current_row, column=col_idx).number_format = '0.000000'
                    ws_results.cell(row=current_row, column=col_idx).font = Font(name='Calibri', size=10)
                    ws_results.cell(row=current_row, column=col_idx).alignment = Alignment(horizontal='right', vertical='center')
                    col_idx += 1

            # G column
            if geo_dict is not None:
                val = geo_dict[state]
                ws_results.cell(row=current_row, column=col_idx, value=float(val))
                ws_results.cell(row=current_row, column=col_idx).number_format = '0.000000'
                ws_results.cell(row=current_row, column=col_idx).font = Font(name='Calibri', size=10)
                ws_results.cell(row=current_row, column=col_idx).alignment = Alignment(horizontal='right', vertical='center')
                col_idx += 1

            # Deployed by column
            if deploying_coalitions is not None:
                deployer = deploying_coalitions.get(state, '')
                ws_results.cell(row=current_row, column=col_idx, value=deployer)
                ws_results.cell(row=current_row, column=col_idx).font = Font(name='Calibri', size=10)
                ws_results.cell(row=current_row, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')

            current_row += 1

        # Highlight max V per player column (amber, matching payoff table style)
        if value_functions is not None:
            _best_fill = PatternFill(start_color='FFFFEEBA', end_color='FFFFEEBA', fill_type='solid')
            for p_idx, player in enumerate(players):
                col_idx = 2 + p_idx
                max_val = float(value_functions[player].max())
                for row_idx, state in enumerate(states):
                    if math.isclose(float(value_functions.loc[state, player]), max_val, rel_tol=1e-13):
                        ws_results.cell(row=3 + row_idx, column=col_idx).fill = _best_fill

        # Add borders to entire table
        table_end_row = current_row - 1
        for r in range(1, table_end_row + 1):
            for c in range(1, num_cols + 1):
                ws_results.cell(row=r, column=c).border = Border(
                    left=thin_border, right=thin_border,
                    top=thin_border, bottom=thin_border
                )

    # Add Short-term Values (static payoffs u) sheet if provided
    if static_payoffs is not None:
        ws_short = wb.create_sheet(title="Short-term Values")

        num_short_cols = 1 + len(players)
        ws_short.column_dimensions['A'].width = 12
        for i in range(1, num_short_cols):
            ws_short.column_dimensions[get_column_letter(1 + i)].width = 12

        # Title
        ws_short.cell(row=1, column=1, value="Short-term Payoffs (u)")
        ws_short.cell(row=1, column=1).font = Font(name='Calibri', bold=True, size=12)
        ws_short.cell(row=1, column=1).fill = PatternFill(
            start_color='FF4D9CC9', end_color='FF4D9CC9', fill_type='solid'
        )
        ws_short.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_short_cols)

        # Headers
        ws_short.cell(row=2, column=1, value="State")
        ws_short.cell(row=2, column=1).font = Font(name='Calibri', bold=True, size=10)
        ws_short.cell(row=2, column=1).fill = PatternFill(
            start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid'
        )
        for idx, player in enumerate(players):
            ws_short.cell(row=2, column=2 + idx, value=player)
            ws_short.cell(row=2, column=2 + idx).font = Font(name='Calibri', bold=True, size=10)
            ws_short.cell(row=2, column=2 + idx).fill = PatternFill(
                start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid'
            )
            ws_short.cell(row=2, column=2 + idx).alignment = Alignment(horizontal='center', vertical='center')

        # Data rows
        for row_idx, state in enumerate(states):
            ws_short.cell(row=3 + row_idx, column=1, value=state)
            ws_short.cell(row=3 + row_idx, column=1).font = Font(name='Calibri', size=10)
            ws_short.cell(row=3 + row_idx, column=1).fill = PatternFill(
                start_color='FFEEBF99', end_color='FFEEBF99', fill_type='solid'
            )
            for p_idx, player in enumerate(players):
                val = float(static_payoffs.loc[state, player])
                ws_short.cell(row=3 + row_idx, column=2 + p_idx, value=val)
                ws_short.cell(row=3 + row_idx, column=2 + p_idx).number_format = '0.000000'
                ws_short.cell(row=3 + row_idx, column=2 + p_idx).font = Font(name='Calibri', size=10)
                ws_short.cell(row=3 + row_idx, column=2 + p_idx).alignment = Alignment(
                    horizontal='right', vertical='center'
                )

        # Highlight top-3 u per player column in descending intensity.
        _rank_fills = [
            PatternFill(start_color='FFFFD966', end_color='FFFFD966', fill_type='solid'),  # 1st: gold
            PatternFill(start_color='FFFFEB9C', end_color='FFFFEB9C', fill_type='solid'),  # 2nd: light gold
            PatternFill(start_color='FFFFF2CC', end_color='FFFFF2CC', fill_type='solid'),  # 3rd: pale gold
        ]
        for p_idx, player in enumerate(players):
            col_idx = 2 + p_idx
            sorted_vals = sorted(static_payoffs[player].dropna().unique(), reverse=True)
            top_vals = sorted_vals[:3]
            for row_idx, state in enumerate(states):
                val = float(static_payoffs.loc[state, player])
                for rank, top_val in enumerate(top_vals):
                    if math.isclose(val, float(top_val), rel_tol=1e-13):
                        ws_short.cell(row=3 + row_idx, column=col_idx).fill = _rank_fills[rank]
                        break

        # Borders
        for r in range(1, 3 + len(states)):
            for c in range(1, num_short_cols + 1):
                ws_short.cell(row=r, column=c).border = Border(
                    left=thin_border, right=thin_border,
                    top=thin_border, bottom=thin_border
                )

    # Add Transition Matrix (P) sheet if provided
    if transition_matrix is not None:
        ws_trans = wb.create_sheet(title="Transition Matrix")

        n_states = len(states)
        ws_trans.column_dimensions['A'].width = 12
        for i in range(1, n_states + 1):
            ws_trans.column_dimensions[get_column_letter(1 + i)].width = 12

        # Title
        ws_trans.cell(row=1, column=1, value="Transition Probability Matrix (P)")
        ws_trans.cell(row=1, column=1).font = Font(name='Calibri', bold=True, size=12)
        ws_trans.cell(row=1, column=1).fill = PatternFill(
            start_color='FF4D9CC9', end_color='FF4D9CC9', fill_type='solid'
        )
        ws_trans.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_states + 1)

        # "From \ To" corner header
        ws_trans.cell(row=2, column=1, value="From \\ To")
        ws_trans.cell(row=2, column=1).font = Font(name='Calibri', bold=True, size=10)
        ws_trans.cell(row=2, column=1).fill = PatternFill(
            start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid'
        )

        # Column headers (target states)
        for j, state in enumerate(states):
            ws_trans.cell(row=2, column=2 + j, value=state)
            ws_trans.cell(row=2, column=2 + j).font = Font(name='Calibri', bold=True, size=10)
            ws_trans.cell(row=2, column=2 + j).fill = PatternFill(
                start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid'
            )
            ws_trans.cell(row=2, column=2 + j).alignment = Alignment(horizontal='center', vertical='center')

        # Data rows
        for i, from_state in enumerate(states):
            ws_trans.cell(row=3 + i, column=1, value=from_state)
            ws_trans.cell(row=3 + i, column=1).font = Font(name='Calibri', bold=True, size=10)
            ws_trans.cell(row=3 + i, column=1).fill = PatternFill(
                start_color='FFEEBF99', end_color='FFEEBF99', fill_type='solid'
            )
            for j, to_state in enumerate(states):
                val = float(transition_matrix.loc[from_state, to_state])
                cell = ws_trans.cell(row=3 + i, column=2 + j, value=val)
                cell.number_format = '0.000000'
                cell.font = Font(name='Calibri', size=10)
                cell.alignment = Alignment(horizontal='right', vertical='center')
                # Highlight diagonal (self-loops / absorbing) in light yellow
                if i == j:
                    cell.fill = PatternFill(
                        start_color='FFFFFF99', end_color='FFFFFF99', fill_type='solid'
                    )
                elif val > 0:
                    cell.fill = PatternFill(
                        start_color='FFCCECE3', end_color='FFCCECE3', fill_type='solid'
                    )

        # Borders
        for r in range(1, 3 + n_states):
            for c in range(1, n_states + 2):
                ws_trans.cell(row=r, column=c).border = Border(
                    left=thin_border, right=thin_border,
                    top=thin_border, bottom=thin_border
                )

    # Add metadata sheet if metadata is provided
    if metadata is not None:
        ws_meta = wb.create_sheet(title="Metadata")

        # Set column widths
        ws_meta.column_dimensions['A'].width = 25
        ws_meta.column_dimensions['B'].width = 40

        # Write header
        ws_meta.cell(row=1, column=1, value="Parameter")
        ws_meta.cell(row=1, column=2, value="Value")
        ws_meta.cell(row=1, column=1).font = Font(name='Calibri', bold=True, size=11)
        ws_meta.cell(row=1, column=2).font = Font(name='Calibri', bold=True, size=11)
        ws_meta.cell(row=1, column=1).fill = PatternFill(
            start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid'
        )
        ws_meta.cell(row=1, column=2).fill = PatternFill(
            start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid'
        )

        # Write metadata rows
        row = 2
        for key, value in metadata.items():
            ws_meta.cell(row=row, column=1, value=str(key))
            ws_meta.cell(row=row, column=2, value=str(value))
            ws_meta.cell(row=row, column=1).font = Font(name='Calibri', size=10)
            ws_meta.cell(row=row, column=2).font = Font(name='Calibri', size=10)
            row += 1

        # Add borders to metadata table
        for r in range(1, row):
            for c in range(1, 3):
                ws_meta.cell(row=r, column=c).border = Border(
                    left=thin_border, right=thin_border,
                    top=thin_border, bottom=thin_border
                )

    # Save workbook
    wb.save(excel_file_path)


def write_payoff_table_excel(
    payoff_df: pd.DataFrame,
    excel_file_path: str,
    players: list[str],
    metadata: dict | None = None,
    source_label: str = "computed",
    sai_column_name: str = "W_SAI_sum_generated",
):
    """
    Write a payoff table Excel file compatible with _load_payoff_table().

    Expected input:
    - payoff_df index: deployer keys (e.g. '( )', '(USA)', '(CHNNDE)')
    - payoff_df columns: one column per player plus optional 'G' column
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Payoffs"

    has_g = "G" in payoff_df.columns
    data_cols = list(players) + ([sai_column_name] if has_g else [])
    headers = ["State"] + data_cols + ["Source file"]

    ws.column_dimensions["A"].width = 18
    for idx in range(len(data_cols)):
        ws.column_dimensions[get_column_letter(2 + idx)].width = 16
    ws.column_dimensions[get_column_letter(2 + len(data_cols))].width = 28

    # Row 1 title so row 2 can be the header for _load_payoff_table(header=1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1, value="Precomputed Payoffs (Generated)").font = Font(
        name="Calibri", bold=True, size=12
    )
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=1, column=1).fill = PatternFill(start_color="FF4D9CC9", end_color="FF4D9CC9", fill_type="solid")

    # Row 2 headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(name="Calibri", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="FFD3D3D3", end_color="FFD3D3D3", fill_type="solid")

    # Data rows
    row = 3
    for state_key in payoff_df.index.tolist():
        ws.cell(row=row, column=1, value=state_key)
        ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="FFEEBF99", end_color="FFEEBF99", fill_type="solid")

        col = 2
        for player in players:
            val = float(payoff_df.loc[state_key, player])
            c = ws.cell(row=row, column=col, value=val)
            c.number_format = "0.000000"
            c.font = Font(name="Calibri", size=10)
            c.alignment = Alignment(horizontal="right", vertical="center")
            col += 1

        if has_g:
            g_val = float(payoff_df.loc[state_key, "G"])
            c = ws.cell(row=row, column=col, value=g_val)
            c.number_format = "0.000"
            c.font = Font(name="Calibri", size=10)
            c.alignment = Alignment(horizontal="right", vertical="center")
            col += 1

        src = ws.cell(row=row, column=col, value=source_label)
        src.font = Font(name="Calibri", italic=True, size=9)
        src.alignment = Alignment(horizontal="left", vertical="center")
        row += 1

    # Highlight top-3 payoffs per player column in descending intensity.
    rank_fills = [
        PatternFill(start_color="FFFFD966", end_color="FFFFD966", fill_type="solid"),  # 1st: gold
        PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid"),  # 2nd: light gold
        PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid"),  # 3rd: pale gold
    ]
    for p_idx, player in enumerate(players):
        col_idx = 2 + p_idx
        sorted_vals = sorted(payoff_df[player].dropna().unique(), reverse=True)
        top_vals = sorted_vals[:3]
        for r_idx, state_key in enumerate(payoff_df.index.tolist(), start=3):
            val = float(payoff_df.loc[state_key, player])
            for rank, top_val in enumerate(top_vals):
                if math.isclose(val, float(top_val), rel_tol=1e-13):
                    ws.cell(row=r_idx, column=col_idx).fill = rank_fills[rank]
                    break

    thin = Side(style="thin", color="000000")
    for r in range(1, row):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = Border(left=thin, right=thin, top=thin, bottom=thin)

    if metadata is not None:
        ws_meta = wb.create_sheet(title="Metadata")
        ws_meta.column_dimensions["A"].width = 28
        ws_meta.column_dimensions["B"].width = 48
        ws_meta.cell(row=1, column=1, value="Parameter").font = Font(name="Calibri", bold=True, size=10)
        ws_meta.cell(row=1, column=2, value="Value").font = Font(name="Calibri", bold=True, size=10)
        ws_meta.cell(row=1, column=1).fill = PatternFill(start_color="FFD3D3D3", end_color="FFD3D3D3", fill_type="solid")
        ws_meta.cell(row=1, column=2).fill = PatternFill(start_color="FFD3D3D3", end_color="FFD3D3D3", fill_type="solid")
        meta_row = 2
        for key, value in metadata.items():
            ws_meta.cell(row=meta_row, column=1, value=str(key))
            ws_meta.cell(row=meta_row, column=2, value=str(value))
            meta_row += 1
        for r in range(1, meta_row):
            for c in range(1, 3):
                ws_meta.cell(row=r, column=c).border = Border(left=thin, right=thin, top=thin, bottom=thin)

    Path(excel_file_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(excel_file_path)


if __name__ == '__main__':
    # Example usage
    import sys
    if len(sys.argv) > 1:
        # Read existing DataFrame and rewrite
        input_file = sys.argv[1]
        output_file = sys.argv[2] if len(sys.argv) > 2 else input_file
        players = sys.argv[3].split(',') if len(sys.argv) > 3 else ['W', 'T', 'C']

        df = pd.read_excel(input_file, header=[0, 1], index_col=[0, 1, 2])
        write_strategy_table_excel(df, output_file, players)
        print(f"Rewrote {input_file} to {output_file}")
