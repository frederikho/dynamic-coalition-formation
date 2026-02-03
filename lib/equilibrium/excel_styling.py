"""
Excel styling utilities for strategy tables.

This module provides functions to apply consistent styling to equilibrium strategy
Excel files to match the format of the original strategy tables.
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import pandas as pd


# Color scheme (matching strategy_tables/weak_governance.xlsx)
COLORS = {
    'proposer_header': 'FF4D9CC9',      # Blue for "Proposer X"
    'state_name': 'FFEEBF99',           # Beige/orange for state names in column A
    'proposition_row': 'FF99C7E0',      # Light blue for proposition rows
    'acceptance_row': 'FFCCECE3',       # Very light blue/green for acceptance rows
    'player_name': 'FF66C5AB',          # Green for player names in acceptance rows
    'self_acceptance': 'FFE28E4D',      # Orange for player accepting own proposal
    'nan_cell': 'FFD9D9D9',             # Gray for NaN/None cells
}

# Column widths (from original file)
COLUMN_WIDTHS = {
    'A': 5.109375,
    'B': 8.77734375,
    'C': 3.77734375,
    # Data columns get width based on pattern
    'data_short': 2.21875,    # For NaN columns
    'data_med': 3.5546875,     # For most data
    'data_long': 4.21875,      # For some data
    'data_xlarge': 5.0,        # For some data
}


def apply_strategy_table_styling(excel_file_path: str, players: list):
    """
    Apply consistent styling to a strategy table Excel file.

    This function matches the styling of the original strategy tables in
    strategy_tables/weak_governance.xlsx and similar files.

    Args:
        excel_file_path: Path to the Excel file to style
        players: List of player names (e.g., ['W', 'T', 'C'])
    """
    # Load workbook
    wb = load_workbook(excel_file_path)
    ws = wb.active

    # Set column widths
    ws.column_dimensions['A'].width = COLUMN_WIDTHS['A']
    ws.column_dimensions['B'].width = COLUMN_WIDTHS['B']
    ws.column_dimensions['C'].width = COLUMN_WIDTHS['C']

    # Data columns use a repeating pattern
    data_widths = [
        COLUMN_WIDTHS['data_short'],
        COLUMN_WIDTHS['data_med'],
        COLUMN_WIDTHS['data_long'],
        COLUMN_WIDTHS['data_xlarge'],
        COLUMN_WIDTHS['data_xlarge'],
    ]

    for col_idx in range(4, 100):  # Plenty of columns
        col_letter = chr(64 + col_idx) if col_idx <= 26 else None
        if col_letter is None:
            break
        width_idx = (col_idx - 4) % len(data_widths)
        ws.column_dimensions[col_letter].width = data_widths[width_idx]

    # Set row heights
    ws.row_dimensions[1].height = 16.5
    ws.row_dimensions[2].height = 14.85
    ws.row_dimensions[3].height = 14.85

    # Get the DataFrame to understand structure
    df = pd.read_excel(excel_file_path, header=[0, 1], index_col=[0, 1, 2])

    # Determine the row/column mapping
    n_states = len(df.index.get_level_values(0).unique())
    n_players = len(players)

    # Apply styling row by row
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for col_idx, cell in enumerate(row, start=1):
            # Row 1: Proposer header
            if row_idx == 1:
                if col_idx >= 4 and cell.value is not None and 'Proposer' in str(cell.value):
                    cell.fill = PatternFill(start_color=COLORS['proposer_header'],
                                          end_color=COLORS['proposer_header'],
                                          fill_type='solid')
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # Row 2: Next State header
            elif row_idx == 2:
                if col_idx >= 4 and cell.value is not None:
                    cell.fill = PatternFill(start_color=COLORS['proposition_row'],
                                          end_color=COLORS['proposition_row'],
                                          fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # Row 3: Index column names (Current State, Type, Player)
            elif row_idx == 3:
                if cell.value is not None:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # Row 4+: Data rows
            elif row_idx >= 4:
                if col_idx == 1:
                    # State name in column A
                    if cell.value is not None and cell.value != 'None':
                        cell.fill = PatternFill(start_color=COLORS['state_name'],
                                              end_color=COLORS['state_name'],
                                              fill_type='solid')
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                elif col_idx == 2:
                    # Type column (Proposition/Acceptance)
                    if cell.value == 'Proposition':
                        cell.fill = PatternFill(start_color=COLORS['proposition_row'],
                                              end_color=COLORS['proposition_row'],
                                              fill_type='solid')
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif cell.value == 'Acceptance':
                        cell.fill = PatternFill(start_color=COLORS['acceptance_row'],
                                              end_color=COLORS['acceptance_row'],
                                              fill_type='solid')
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                elif col_idx == 3:
                    # Player name column
                    if cell.value in players:
                        cell.fill = PatternFill(start_color=COLORS['player_name'],
                                              end_color=COLORS['player_name'],
                                              fill_type='solid')
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                elif col_idx >= 4:
                    # Data cells
                    # Need to find row type - might be in current row or above (merged cells)
                    row_type = ws.cell(row=row_idx, column=2).value
                    if row_type is None or row_type == 'None':
                        # Look upward for the row type
                        for search_row in range(row_idx - 1, 3, -1):
                            row_type = ws.cell(row=search_row, column=2).value
                            if row_type in ['Proposition', 'Acceptance']:
                                break

                    if row_type == 'Proposition':
                        # Proposition row cells - all get light blue
                        cell.fill = PatternFill(start_color=COLORS['proposition_row'],
                                              end_color=COLORS['proposition_row'],
                                              fill_type='solid')
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                    elif row_type == 'Acceptance':
                        # Acceptance row cells
                        # Find player name - might be in current row or above
                        player_name = ws.cell(row=row_idx, column=3).value
                        if player_name is None or player_name == 'None' or player_name not in players:
                            # Look upward for player name
                            for search_row in range(row_idx - 1, 3, -1):
                                player_name = ws.cell(row=search_row, column=3).value
                                if player_name in players:
                                    break

                        # Determine which proposer this column belongs to
                        proposer_name = ws.cell(row=1, column=col_idx).value
                        if proposer_name and 'Proposer' in str(proposer_name):
                            proposer = proposer_name.split()[-1]  # Extract player name

                            # Check if this is a self-acceptance (player accepting own proposal)
                            is_self_acceptance = (player_name == proposer)

                            if cell.value is None or (isinstance(cell.value, str) and cell.value in ['None', 'nan']):
                                # NaN cells
                                cell.fill = PatternFill(start_color=COLORS['nan_cell'],
                                                      end_color=COLORS['nan_cell'],
                                                      fill_type='solid')
                            elif is_self_acceptance:
                                # Self-acceptance (player accepting their own proposal)
                                cell.fill = PatternFill(start_color=COLORS['self_acceptance'],
                                                      end_color=COLORS['self_acceptance'],
                                                      fill_type='solid')
                            else:
                                # Regular acceptance
                                cell.fill = PatternFill(start_color=COLORS['acceptance_row'],
                                                      end_color=COLORS['acceptance_row'],
                                                      fill_type='solid')

                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        else:
                            # Fallback
                            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Save the styled workbook
    wb.save(excel_file_path)


if __name__ == '__main__':
    # Example usage
    import sys
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
        players = ['W', 'T', 'C']  # Default players
        if len(sys.argv) > 2:
            players = sys.argv[2].split(',')
        apply_strategy_table_styling(excel_file, players)
        print(f"Applied styling to {excel_file}")
    else:
        print("Usage: python excel_styling.py <excel_file> [players]")
        print("Example: python excel_styling.py output.xlsx W,T,C")
